from __future__ import annotations

import calendar
from copy import copy
from datetime import date, time
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_TEMPLATE_PATH = Path(getattr(settings, "BASE_DIR", Path("."))) / "api" / "assets" / "Modelo.xlsx"

PT_BR_WEEKDAYS = [
    "segunda-feira",
    "terça-feira",
    "quarta-feira",
    "quinta-feira",
    "sexta-feira",
    "sábado",
    "domingo",
]

PT_BR_MONTHS = [
    "janeiro",
    "fevereiro",
    "março",
    "abril",
    "maio",
    "junho",
    "julho",
    "agosto",
    "setembro",
    "outubro",
    "novembro",
    "dezembro",
]


def format_data_extenso(d: date) -> str:
    return f"{PT_BR_WEEKDAYS[d.weekday()]}, {d.day} de {PT_BR_MONTHS[d.month - 1]} de {d.year}"


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _time_for_sort(t: Optional[time]) -> time:
    return t if t is not None else time(23, 59, 59)


def _choose_value(entries: List[Dict[str, Any]], field: str) -> Any:
    """
    Regra:
      - Se NÃO houver divergência (<=1 valor não-vazio distinto): pega a primeira entrada (menor ordem) com valor
      - Se houver divergência: pega a última entrada (maior ordem) que esteja preenchida
    """
    non_empty = [e.get(field) for e in entries if not _is_blank(e.get(field))]
    uniq: List[Any] = []
    for v in non_empty:
        if v not in uniq:
            uniq.append(v)

    if len(uniq) <= 1:
        for e in sorted(entries, key=lambda x: int(x.get("ordem") or 0)):
            v = e.get(field)
            if not _is_blank(v):
                return v
        return None

    for e in sorted(entries, key=lambda x: int(x.get("ordem") or 0), reverse=True):
        v = e.get(field)
        if not _is_blank(v):
            return v
    return None


def _clone_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 12) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):  # A=1 ... L=12
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)

        dst.value = None


def _ensure_rows(ws: Worksheet, total_lines: int, start_row: int = 11, base_end_row: int = 40) -> None:
    base_capacity = base_end_row - start_row + 1  # 30
    if total_lines <= base_capacity:
        return

    extra = total_lines - base_capacity
    insert_at = base_end_row + 1
    ws.insert_rows(insert_at, amount=extra)

    for i in range(extra):
        _clone_row_style(ws, base_end_row, insert_at + i, max_col=12)


def _clear_body(ws: Worksheet, start_row: int = 11, min_end_row: int = 40) -> None:
    end_row = max(ws.max_row, min_end_row)
    for r in range(start_row, end_row + 1):
        for c in range(1, 13):  # A..L
            ws.cell(row=r, column=c).value = None


def _build_lines_by_day(rows: List[Dict[str, Any]]) -> Dict[int, List[Dict[str, Any]]]:
    sessions: Dict[int, Dict[str, Any]] = {}

    for r in rows:
        rid = int(r["registro_id"])
        sess = sessions.setdefault(
            rid,
            {
                "registro_id": rid,
                "data": r["data"],
                "sala_nome": r["sala_nome"],
                "em_aberto": bool(r.get("em_aberto")),
                "rows": [],
            },
        )
        sess["rows"].append(r)

    lines_by_day: Dict[int, List[Dict[str, Any]]] = {}

    for sess in sessions.values():
        raw_rows: List[Dict[str, Any]] = sess["rows"]

        # 1) Colapsa por ordem (pega a “última” por (seq, entrada_id))
        by_ordem: Dict[int, Dict[str, Any]] = {}
        for rr in raw_rows:
            ordem = int(rr.get("ordem") or 0)
            cur = by_ordem.get(ordem)
            rr_key = (int(rr.get("seq") or 0), int(rr.get("entrada_id") or 0))
            if cur is None:
                by_ordem[ordem] = rr
            else:
                cur_key = (int(cur.get("seq") or 0), int(cur.get("entrada_id") or 0))
                if rr_key > cur_key:
                    by_ordem[ordem] = rr

        if not by_ordem:
            continue

        # 2) FIM (G) por sessão: pega qualquer horario_termino preenchido (se existir)
        fim_candidates = [rr.get("horario_termino") for rr in raw_rows if rr.get("horario_termino")]
        fim = max(fim_candidates) if fim_candidates else None

        is_open = bool(sess.get("em_aberto")) or fim is None
        fim_out = None if is_open else fim
        obs = "Evento não encerrado" if is_open else ""

        max_ordem = max(by_ordem.keys())
        num_groups = (max_ordem + 2) // 3  # grupos de 3: 1-3, 4-6, ...

        for g in range(num_groups):
            start = g * 3 + 1
            group_entries = [by_ordem[o] for o in range(start, start + 3) if o in by_ordem]
            if not group_entries:
                continue

            group_entries_sorted = sorted(group_entries, key=lambda x: int(x.get("ordem") or 0))

            comissao_nome = _choose_value(group_entries_sorted, "comissao_nome")
            atividade = None
            if comissao_nome:
                atividade = comissao_nome.split("-")[0].strip()

            line = {
                "registro_id": sess["registro_id"],
                "group_index": g,
                "data": sess["data"],
                "sala_nome": sess["sala_nome"],
                "atividade_legislativa": atividade,
                "nome_evento": _choose_value(group_entries_sorted, "nome_evento"),
                "horario_pauta": _choose_value(group_entries_sorted, "horario_pauta"),
                "horario_inicio": _choose_value(group_entries_sorted, "horario_inicio"),
                "horario_termino": fim_out,
                "op1": (by_ordem.get(start) or {}).get("operador_nome_exibicao"),
                "op2": (by_ordem.get(start + 1) or {}).get("operador_nome_exibicao"),
                "op3": (by_ordem.get(start + 2) or {}).get("operador_nome_exibicao"),
                "obs": obs,
            }

            day = int(sess["data"].day)
            lines_by_day.setdefault(day, []).append(line)

    # ordenação dentro do dia:
    # 1) horario_pauta (asc)
    # 2) horario_inicio (asc)
    # 3) cadastro.sala.nome (asc)
    for day, lines in lines_by_day.items():
        lines.sort(
            key=lambda l: (
                _time_for_sort(l.get("horario_pauta")),
                _time_for_sort(l.get("horario_inicio")),
                (l.get("sala_nome") or ""),
                int(l.get("registro_id") or 0),
                int(l.get("group_index") or 0),
            )
        )

    return lines_by_day


def gerar_rds_xlsx(
    ano: int,
    mes: int,
    rows: List[Dict[str, Any]],
    template_path: Optional[Path] = None,
) -> bytes:
    template = template_path or DEFAULT_TEMPLATE_PATH
    if not template.exists():
        raise FileNotFoundError(f"Template não encontrado: {template}")

    wb = load_workbook(template)

    last_day = calendar.monthrange(ano, mes)[1]
    lines_by_day = _build_lines_by_day(rows)

    for d in range(1, 32):
        sheet_name = f"{d:02d}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # limpa corpo sempre (pra garantir)
        _clear_body(ws)

        # dias inexistentes (ex.: 31 em abril) => A7 em branco e corpo vazio
        if d > last_day:
            ws["A7"].value = None
            continue

        lines = lines_by_day.get(d) or []

        # dias existentes, mas sem registros => A7 em branco e corpo vazio
        if not lines:
            ws["A7"].value = None
            continue

        # dia com registros => preenche A7 e tabela
        ws["A7"].value = format_data_extenso(date(ano, mes, d))

        _ensure_rows(ws, total_lines=len(lines))

        for idx, line in enumerate(lines):
            row = 11 + idx

            ws[f"A{row}"].value = line.get("sala_nome") or None
            ws[f"B{row}"].value = "SGM"
            ws[f"C{row}"].value = line.get("atividade_legislativa") or None
            ws[f"D{row}"].value = (line.get("nome_evento") or None)
            ws[f"E{row}"].value = line.get("horario_pauta") or None
            ws[f"F{row}"].value = None
            ws[f"G{row}"].value = line.get("horario_inicio") or None
            ws[f"H{row}"].value = line.get("horario_termino") or None
            ws[f"I{row}"].value = (line.get("op1") or None)
            ws[f"J{row}"].value = (line.get("op2") or None)
            ws[f"K{row}"].value = (line.get("op3") or None)
            ws[f"L{row}"].value = (line.get("obs") or None)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()